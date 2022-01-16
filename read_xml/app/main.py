import sqlite3
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook

from settings import *


def format_key(tag: str):
    text = tag.split('}')[-1].lower()
    return text


def get_count(key: str, counts: dict):
    if key in counts:
        value = counts[key]
        value.append('1')
    else:
        counts[key] = ['1']
    count = len(counts[key])
    return count


def write_text_file(file_name='output.txt', file_text=''):
    log.debug(f'Сохранение текстового файла {file_name}')
    with open(file_name, 'w', encoding="utf-8") as file:
        file.write(file_text)
    log.info(f'Текстовый файл {file_name} сохранён')


class ManageSave:
    def __init__(self, folder_name, data, tags):
        self.folder_name = folder_name
        self.data = data
        self.tags = tags
        self.df = None
        if folder_name == '01':
            self.tags_organisation = {'organisation_main_crit_rate_value',
                                      'organisation_summary_all_values_types',
                                      'organisation_summary_main_values_types',
                                      'organization_code',
                                      'organization_name', }
            self.check_tags = {'full_name', 'short_name', 'ppo_name',
                               'opf_name', 'tofk_name', 'council_name',
                               'summary_results', 'organization_name'}
            self.tag_organisation = 'organisations'
        elif folder_name == '02':
            self.tags_organisation = {'org_group_code', 'org_group_name', }
            self.check_tags = {'full_name', 'short_name', 'opf_name',
                               'org_group_name', }
            self.tag_organisation = 'org_groups'
        elif folder_name == '03':
            self.tags_organisation = {'name', 'date', 'url', 'type', }
            self.check_tags = {'fullname', 'main_okfs_name', 'main_okfs_name',
                               'main_okfs_name', 'main_okfs_name',
                               'shortname', }
            self.tag_organisation = 'documents'
        elif folder_name == '04':
            self.tags_organisation = {'organization_code', 'organization_name',
                                      'review_info', 'quality_improve_plans', }
            self.check_tags = {'organization_name', 'full_name', 'short_name',
                               'scope_name', 'council_name', }
            self.tag_organisation = 'organisations'
        self.all_tags = self.tags | self.tags_organisation

    def set_pandas_dat(self, data):
        columns = list(self.all_tags)
        columns.sort()
        self.df = pd.DataFrame(data, columns=columns)

    def check_regions(self, data: dict):
        for key in self.check_tags:
            try:
                text = str(data[key]).lower()
            except Exception as e:
                text = ''

            if 'саха' in text or 'якутия' in text:
                return True
        return False

    def create_date(self):
        data = {}

        for tag in self.all_tags:
            data[tag] = []

        if self.data[self.tag_organisation]:
            for organisation in self.data[self.tag_organisation]:
                for tag in self.tags:
                    # organisation
                    if tag in self.tags_organisation:
                        try:
                            data[tag].append(organisation[tag])
                        except Exception as e:
                            data[tag].append('')

                    else:
                        # other tags
                        try:
                            data[tag].append(self.data[tag])
                        except Exception as e:
                            data[tag].append('')

        else:
            for tag in self.all_tags:
                try:
                    data[tag].append(self.data[tag])
                except Exception as e:
                    data[tag].append('None')

        flag_region = self.check_regions(self.data)
        return data, flag_region

    def get_file_name(self):
        if self.folder_name == '01':
            file_name = self.data['agency_inn']
        elif self.folder_name == '02':
            file_name = self.data['inn']
        elif self.folder_name == '03':
            file_name = self.data['inn']
        elif self.folder_name == '04':
            file_name = self.data['agency_inn']
        else:
            file_name = ''
        return file_name

    def manage_save(self):
        data, flag_region = self.create_date()
        file_name = self.get_file_name()
        self.set_pandas_dat(data=data)

        self.write_csv(f'_all')
        self.write_db(f'_all')
        self.write_xls(file_name)
        self.write_db(file_name)
        if flag_region:
            self.write_csv('_якутия')
            self.write_db('_якутия')

    def write_csv(self, file_name_):
        file_name = f'{DATA}{self.folder_name}/{file_name_}.csv'
        self.df.to_csv(file_name, index=False, header=False, mode='a')

    def write_xls(self, file_name_):
        file_name = f'{DATA}{self.folder_name}/{file_name_}.xlsx'
        try:
            book = load_workbook(file_name)
            writer = pd.ExcelWriter(file_name, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            for sheet_name in writer.sheets:
                self.df.to_excel(writer,
                                 sheet_name=sheet_name,
                                 startrow=writer.sheets[sheet_name].max_row,
                                 index=False,
                                 header=False)
            writer.save()
        except Exception as e:
            self.df.to_excel(file_name, index=False)

    def write_db(self, file_name_):
        file_name = f'{DATA}{self.folder_name}/{file_name_}.db'
        table_name = f'table_{self.folder_name}'
        conn = sqlite3.connect(file_name)
        c = conn.cursor()
        command = f"CREATE TABLE IF NOT EXISTS {table_name} ({','.join(self.all_tags)})"
        c.execute(command)
        conn.commit()
        try:
            self.df.to_sql(table_name, conn, if_exists='append', index=False)
        except Exception as e:
            pass


class ParserXML:
    def __init__(self, xml_file, folder_name):
        self.xml_file = xml_file
        self.folder_name = folder_name
        self.data = {}
        self.tags = self.tags()
        log.debug(f'read file {xml_file}')

    def tags(self):
        if self.folder_name == '01':
            tags = {'id', 'full_name', 'dop_crit_rate_values', 'short_name',
                    'ppo_code', 'ppo_name',
                    'opf_code', 'opf_name', 'agency_inn', 'agency_kpp',
                    'tofk_code', 'tofk_name',
                    'sr_code', 'gmy_code', 'create_date', 'scope_code',
                    'scope_name', 'council_code',
                    'council_name', 'doc_type', 'doc_date', 'doc_num',
                    'summary_results', }
        elif self.folder_name == '02':
            tags = {'full_name', 'short_name', 'opf_code', 'opf_name',
                    'okfs_code', 'okfs_name',
                    'inn', 'kpp', 'reg_date', 'scope_code', 'scope_name',
                    'email', 'phone_num',
                    'rating_year', 'subject_code', 'subject_name',
                    'country_code', 'country_name',
                    'zip', 'region_name', 'rajon_name', 'street_name',
                    'building', 'oktmo_code',
                    'oktmo_name', }
        elif self.folder_name == '03':
            tags = {'positionid', 'changedate', 'main_okato_code', 'regnum',
                    'fullname', 'inn', 'kpp',
                    'shortname', 'ogrn', 'orgtype', 'special', 'email', 'phone',
                    'www', }
        elif self.folder_name == '04':
            tags = {'id', 'full_name', 'short_name', 'agency_inn', 'agency_kpp',
                    'sr_code', 'gmy_code',
                    'review_date', 'review_year', 'scope_code', 'scope_name',
                    'result_summary', 'file_name',
                    'file_size', 'upload_date', 'council_code', 'council_name',
                    'quality_improve_plans', }

        return tags

    def manage_parse(self):
        if self.folder_name == '01':
            self.parse_xml_01()
        elif self.folder_name == '02':
            self.parse_xml_02()
        elif self.folder_name == '03':
            self.parse_xml_03()
        elif self.folder_name == '04':
            self.parse_xml_04()
        return self.data, self.tags

    def parse_xml_01(self):
        self.data['organisations'] = []
        for (ev, el) in ET.iterparse(self.xml_file):
            # all tags
            if format_key(tag=el.tag) in self.tags:
                self.data[format_key(el.tag)] = el.text

            # summary
            if format_key(tag=el.tag) == 'rate_value_summary':
                for el_1 in el:
                    if format_key(tag=el_1.tag) == 'summary_value':
                        self.data[format_key(el_1.tag)] = el_1.text
                        self.tags.add(format_key(el_1.tag))
                    elif format_key(tag=el_1.tag) == 'main_crit_rate_values':
                        for el_2 in el_1:
                            self.data[format_key(el_2.tag)] = el_2.text
                            if format_key(el_2.tag) != 'main_crit_rate_value':
                                self.tags.add(format_key(el_2.tag))

            # organisation
            if format_key(tag=el.tag) == 'rate_value_organizations':
                for el_1 in list(el):

                    if format_key(el_1.tag) == 'organization':
                        organisation = {}
                        for el_2 in el_1:
                            key = format_key(el_2.tag)
                            value = el_2.text
                            if key == 'main_crit_rate_values':
                                for el_3 in el_2:
                                    organisation[
                                        format_key(el_3.tag)] = el_3.text
                                    self.tags.add(
                                        'organisation_' + format_key(el_3.tag))
                                continue
                            organisation[key] = value
                            self.tags.add(key)
                        self.data['organisations'].append(organisation)

    def parse_xml_02(self):
        self.data['org_groups'] = []
        for (ev, el) in ET.iterparse(self.xml_file):
            # all tags
            if format_key(tag=el.tag) in self.tags:
                self.data[format_key(el.tag)] = el.text

            # organisation
            if format_key(tag=el.tag) == 'org_groups':
                for el_1 in list(el):

                    if format_key(el_1.tag) == 'org_group':
                        organisation = {}
                        for el_2 in el_1:
                            key = format_key(el_2.tag)
                            value = el_2.text
                            organisation[key] = value
                            self.tags.add(key)
                        self.data['org_groups'].append(organisation)

    def parse_xml_03(self):
        self.data['documents'] = []
        for (ev, el) in ET.iterparse(self.xml_file):
            # all tags
            if format_key(tag=el.tag) in self.tags:
                self.data[format_key(el.tag)] = el.text

            # main
            if format_key(tag=el.tag) == 'main':
                for el_1 in el:
                    if format_key(tag=el_1.tag) == 'ogrn':
                        self.data['main_' + format_key(el_1.tag)] = el_1.text
                        self.tags.add('main_' + format_key(el_1.tag))
                    elif format_key(tag=el_1.tag) == 'classifier':
                        for el_2 in el_1:
                            if format_key(tag=el_2.tag) == 'okpo':
                                self.data[
                                    'main_' + format_key(el_2.tag)] = el_2.text
                                self.tags.add('main_' + format_key(el_2.tag))
                            else:
                                for el_3 in el_2:
                                    self.data[
                                        f'main_{format_key(el_2.tag)}_' +
                                        f'{format_key(el_3.tag)}'] = el_3.text
                                    self.tags.add(
                                        f'main_{format_key(el_2.tag)}_'
                                        f'{format_key(el_3.tag)}')

            # documents
            if format_key(tag=el.tag) == 'document':
                documents = {}
                for el_1 in list(el):
                    key = format_key(el_1.tag)
                    value = el_1.text
                    documents[key] = value
                    self.tags.add(key)
                self.data['documents'].append(documents)

    def parse_xml_04(self):
        self.data['organisations'] = []
        for (ev, el) in ET.iterparse(self.xml_file):
            # all tags
            if format_key(tag=el.tag) in self.tags:
                self.data[format_key(el.tag)] = el.text

            # organisation
            if format_key(tag=el.tag) == 'review_result':
                for el_1 in list(el):
                    if format_key(el_1.tag) == 'review_result_organizations':
                        organisation = {}
                        for el_2 in el_1:
                            key = format_key(el_2.tag)
                            value = el_2.text
                            if key == 'organization':
                                for el_3 in el_2:
                                    organisation[
                                        format_key(el_3.tag)] = el_3.text
                                    self.tags.add(format_key(el_3.tag))
                                continue
                            organisation[key] = value
                            self.tags.add(key)
                        self.data['organisations'].append(organisation)


def main():
    folders = ['01', '02', '03', '04']
    for folder in folders:
        os.chdir(f'{DIR_XML}/{folder}')
        file_names = os.listdir()

        for i, file_name in enumerate(file_names):
            file_name = f'{DIR_XML}{folder}/{file_name}'

            file_xml = ParserXML(xml_file=file_name, folder_name=folder)
            data, tags = file_xml.manage_parse()
            save_module = ManageSave(folder_name=folder, data=data, tags=tags)
            save_module.manage_save()
            if i % 10 == 0:
                log.info(f'Write data #{i + 10} of {len(file_names)}')


if __name__ == '__main__':
    main()
